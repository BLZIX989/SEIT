"""Workbook engine (spec section 29): builds/updates the Master
Calculation Workbook with the required sheets. No pre-existing workbook
was found in the repository (audited at build time), so this is the
first version; future runs must preserve historical sheets rather than
overwrite them wholesale.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

REQUIRED_SHEETS = [
    "FOUNDATION", "OBJECT_REGISTRY", "TYPE_REGISTRY", "TRANSFORMATION_REGISTRY",
    "EQUATION_REGISTRY", "DEPENDENCY_GRAPH", "ASSUMPTIONS", "PROOFS", "CALCULATIONS",
    "VERIFICATION", "FALSIFICATION", "PROVENANCE", "TARGET_INDEPENDENCE",
    "PREDICTIONS", "OPEN_PROBLEMS", "STATUS_MATRIX",
]


def _write_table(ws, rows: list[dict], headers: list[str] | None = None) -> None:
    if not rows:
        ws.append(["(no rows registered in this build)"])
        return
    headers = headers or list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_stringify(row.get(h)) for h in headers])
    for i, h in enumerate(headers, start=1):
        width = max(12, min(60, len(h) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width


def _stringify(v):
    if isinstance(v, (list, dict)):
        return str(v)[:500]
    return v


def build_workbook(result: dict, out_path: Path) -> None:
    registries = result["registries"]
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("FOUNDATION")
    _write_table(ws, [
        {"item": "Logic", "value": "classical propositional/predicate logic (unspecified beyond parameterization)"},
        {"item": "Membership (in)", "value": "parameterizable; not hard-coded to ZFC (spec section 7)"},
        {"item": "F0", "value": "(Logic, in, Axioms) -- parameterizable formal foundation, uninstantiated"},
        {"item": "F1 = EMPTYSET", "value": "constructed as the first object of the forward chain template; OPEN"},
        {"item": "note", "value": "no specific axiom set is asserted as THE physical foundation in this build"},
    ], headers=["item", "value"])

    ws = wb.create_sheet("TYPE_REGISTRY")
    _write_table(ws, registries.types.to_list())

    ws = wb.create_sheet("OBJECT_REGISTRY")
    _write_table(ws, [
        {"id": o["id"], "type": o["type"], "status": o["status"],
         "dependencies": o["dependencies"], "role": o.get("role")}
        for o in registries.objects.to_list()
    ])

    ws = wb.create_sheet("TRANSFORMATION_REGISTRY")
    _write_table(ws, [
        {"id": t["id"], "domain": t["domain"], "codomain": t["codomain"],
         "action": t["action"], "status": t["status"], "dependencies": t["dependencies"]}
        for t in registries.transformations.to_list()
    ])

    ws = wb.create_sheet("EQUATION_REGISTRY")
    _write_table(ws, [
        {"id": e["id"], "lhs": e["lhs"], "rhs": e["rhs"], "domain": e["domain"],
         "status": e["status"], "dependencies": e["dependencies"]}
        for e in registries.equations.to_list()
    ])

    ws = wb.create_sheet("DEPENDENCY_GRAPH")
    edges = []
    for node in registries.all_nodes():
        for dep in node.dependencies:
            edges.append({"node": node.id, "depends_on": dep})
    _write_table(ws, edges)

    ws = wb.create_sheet("ASSUMPTIONS")
    assumptions_rows = []
    for node in registries.all_nodes():
        for a in node.assumptions:
            assumptions_rows.append({"node_id": node.id, "assumption": a})
    _write_table(ws, assumptions_rows)

    ws = wb.create_sheet("PROOFS")
    proofs = [{"id": f"PROOF-{t['id']}", "transformation_id": t["id"],
               "statement": t["action"], "method": t["proof"], "status": t["status"]}
              for t in registries.transformations.to_list() if t.get("proof")]
    _write_table(ws, proofs)

    ws = wb.create_sheet("CALCULATIONS")
    _write_table(ws, [
        {"id": c["id"], "kind": c["kind"], "inputs": c["inputs"], "status": c["status"]}
        for c in result["test_results"]["calculations"]
    ])

    ws = wb.create_sheet("VERIFICATION")
    verif_rows = []
    for c in result["test_results"]["calculations"]:
        verif_rows.append({"calculation_id": c["id"], "verification": c["verification"], "status": c["status"]})
    _write_table(ws, verif_rows)

    ws = wb.create_sheet("FALSIFICATION")
    _write_table(ws, [f.to_dict() for f in result["falsifications"]])

    ws = wb.create_sheet("PROVENANCE")
    prov_rows = []
    for node in registries.all_nodes():
        if node.provenance:
            p = node.provenance
            prov_rows.append({"node_id": node.id, "source": p.source, "status": p.status,
                               "git_commit": p.git_commit[:12], "timestamp": p.execution_timestamp})
    _write_table(ws, prov_rows)

    ws = wb.create_sheet("TARGET_INDEPENDENCE")
    from compiler.falsification.target_independence import scan_registries
    findings = [f.to_dict() for f in scan_registries(registries)]
    _write_table(ws, findings)

    ws = wb.create_sheet("PREDICTIONS")
    _write_table(ws, [
        {"note": "No physical predictions are registered in this build. The gauge, matter, "
                 "thermodynamic, and cosmological engines are gated behind this compiler's "
                 "own self-audit (spec section 41) and have not been activated. A fitted "
                 "parameter is never registered here as a prediction (spec section 4)."}
    ], headers=["note"])

    ws = wb.create_sheet("OPEN_PROBLEMS")
    open_rows = []
    for node in registries.all_nodes():
        if node.status.value == "OPEN":
            open_rows.append({"id": node.id, "kind": type(node).__name__,
                               "note": (node.carrier if hasattr(node, "carrier") and isinstance(node.carrier, str)
                                        else getattr(node, "action", ""))[:300]})
    _write_table(ws, open_rows)

    ws = wb.create_sheet("STATUS_MATRIX")
    _write_table(ws, registries.status_matrix())

    wb.save(out_path)
